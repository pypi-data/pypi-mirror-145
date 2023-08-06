from collections import defaultdict
from copy import copy
import numpy as np
from openpyxl.styles import Color, PatternFill
from pdf2image import convert_from_path
import pathlib
import subprocess
from tempfile import TemporaryDirectory

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .types import BBox, traverse, gather_tables


def write_wb_to_pdf_path(wb: Workbook, path: pathlib.PosixPath):
    with TemporaryDirectory() as d:
        xls_path = pathlib.Path(d).joinpath(path.name.rsplit(".", -1)[0])
        wb.save(xls_path)

        for i in range(10):
            # Now convert it to a PDF
            response = subprocess.run(
                [
                    "soffice",
                    "--convert-to",
                    "pdf",
                    str(xls_path),
                    "--outdir",
                    str(path.parent),
                ],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
            )
            if response.returncode == 0:
                break
        else:
            assert False, "Failed to run soffice: %s" % (response.stderr)


def clone_wb(wb: Workbook):
    with TemporaryDirectory() as d:
        xls_path = pathlib.Path(d).joinpath("foo.xlsx")
        wb.save(xls_path)
        return load_workbook(xls_path)


def write_cells(ws: Worksheet, record):
    def write_record(record):
        ws.cell(row=record.cell.row, column=record.cell.column).value = record.fmt()

    traverse(record, write_record)


def copy_style_range(
    ws: Worksheet,
    src_row_start: int,
    src_col_start: int,
    src_height: int,
    src_width: int,
    dst_row_start: int,
    dst_col_start: int,
):
    for i in range(src_height):
        for j in range(src_width):
            src_cell = ws.cell(src_row_start + i, src_col_start + j)
            if src_cell.has_style:
                # https://stackoverflow.com/questions/23332259/copy-cell-style-openpyxl
                ws.cell(dst_row_start + i, dst_col_start + j)._style = copy(
                    src_cell._style
                )


def rgb(color):
    return (color >> 16) & 0xFF, (color >> 8) & 0xFF, color & 0xFF


def color_cells(ws: Worksheet, record):
    white = 0xFFFFFF
    counter = 1
    colors = {}

    def write_color(record):
        nonlocal colors, counter

        color = white - counter * 19
        counter += 1

        assert color not in colors
        colors[color] = record

        r, g, b = rgb(color)
        assert not (r == g and g == b), "No grayscale allowed"

        # https://stackoverflow.com/questions/30484220/fill-cells-with-colors-using-openpyxl
        ws.cell(row=record.cell.row, column=record.cell.column).fill = PatternFill(
            patternType="solid",
            fill_type="solid",
            fgColor=Color("{:06X}".format(color)),
        )

    traverse(record, write_color)
    return colors


# NOTE: This assumes the workbook has exactly one sheet
def annotate_locations(wb: Workbook, record):
    # Now, color the cells and generate another PDF file that we can use to compute bounding boxes
    colors = color_cells(wb[wb.sheetnames[0]], record)

    with TemporaryDirectory() as d:
        pdf_path = pathlib.Path(d).joinpath("colors.pdf")
        write_wb_to_pdf_path(wb, pdf_path)
        write_wb_to_pdf_path(wb, pathlib.Path("/Users/ankur/pooja.pdf"))

        # https://stackoverflow.com/questions/60051941/find-the-coordinates-in-an-image-where-a-specified-colour-is-detected
        pages = [np.array(p.convert("RGB")) for p in convert_from_path(pdf_path)]

    x_colors = []
    y_colors = []

    uniques = []
    for n, X in enumerate(pages):
        if X.ndim != 3:
            raise ValueError(f"invalid {X.ndim = }")
        num_rows, num_cols, depth = X.shape
        if depth != 3:
            raise ValueError(f"invalid {depth = }")
        x = X.reshape(-1, 3).astype(np.uint32)
        y = x[:, 0] * (256 * 256) + x[:, 1] * 256 + x[:, 2]
        unique_y, inv_indices = np.unique(y, return_inverse=True)
        uniques.append((unique_y, inv_indices))

    for color, value in colors.items():
        found = False
        for n, (X, (unique_y, inv_indices)) in enumerate(zip(pages, uniques)):
            i = np.flatnonzero(unique_y == color)
            if i.size == 0:
                continue

            found = True
            assert i.shape == (1,)
            j = np.flatnonzero(inv_indices == i[0])
            rows, cols = divmod(j, X.shape[1])
            min_row, max_row = rows.min(), rows.max()
            min_col, max_col = cols.min(), cols.max()

            value.location = BBox(
                **{
                    "page": n,
                    "top": min_row / num_rows,
                    "left": min_col / num_cols,
                    "height": (max_row - min_row) / num_rows,
                    "width": (max_col - min_col) / num_cols,
                }
            )
            break
        assert found, "Could not find color 0x%.6X" % (color)


def remove_multipage_rows(wb, record):
    ws = wb[wb.sheetnames[0]]
    for table in gather_tables(record):
        to_remove = []
        for i, row in enumerate(table):
            pages = set()

            def capture_page(record):
                pages.add(record.location.page)

            traverse(row, capture_page)

            if len(pages) > 1:
                to_remove.append(i)

                def clear_cell(record):
                    ws.cell(row=record.cell.row, column=record.cell.column).value = None

                traverse(row, clear_cell)

        for i in reversed(to_remove):
            del table[i]
