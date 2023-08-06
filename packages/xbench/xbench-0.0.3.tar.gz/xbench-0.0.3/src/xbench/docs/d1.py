from copy import copy
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pathlib
from pydantic import BaseModel, validate_arguments
import time
from typing import List

from .. import fmt
from ..types import (
    Cell,
    Config,
    GenOpts,
    NumberLabel,
    TextLabel,
    TimestampLabel,
)
from ..utils import text_dims
from ..xl import (
    annotate_locations,
    clone_wb,
    write_wb_to_pdf_path,
    write_cells,
    remove_multipage_rows,
)
from .doc_type import DocType


class D1Row(BaseModel):
    upc: NumberLabel
    product_description: TextLabel
    site: NumberLabel
    zone: NumberLabel
    date_range: TextLabel
    scan_dollars: NumberLabel
    units: NumberLabel
    amount_dollars: NumberLabel
    scan_deal: NumberLabel


class D1Record(BaseModel):
    company: TextLabel
    addr: TextLabel
    city_state: TextLabel
    inv_date: TimestampLabel
    inv_number: TextLabel

    rows: List[D1Row]


class D1(DocType):
    DOC_NUMBER = 1
    Record = D1Record

    FIRST_TABLE_ROW = 15
    FIRST_TABLE_COLUMN = 1

    @classmethod
    @validate_arguments
    def generate_record(cls, opts: GenOpts) -> D1Record:
        fake = DocType.faker()

        def translate(row_index, col_index, **kwargs):
            return Cell(
                row=cls.FIRST_TABLE_ROW + row_index,
                column=cls.FIRST_TABLE_COLUMN + col_index,
                **kwargs
            )

        addr, city_state = fake.address().split("\n")
        return D1Record(
            company=TextLabel(value=fake.company(), cell=Cell(row=8, column=1)),
            addr=TextLabel(value=addr, cell=Cell(row=9, column=1)),
            city_state=TextLabel(value=city_state, cell=Cell(row=10, column=1)),
            inv_date=TimestampLabel(
                value=datetime.combine(
                    fake.date_between(start_date=DocType.START_DATE),
                    datetime.min.time(),
                ),
                cell=Cell(row=3, column=5),
            ),
            inv_number=TextLabel(
                value=fake.bothify(text="???######", letters="ABCDEFGH"),
                cell=Cell(row=5, column=5),
            ),
            rows=[
                D1Row(
                    upc=TextLabel(
                        value=opts.maybe_empty(fake.bothify(text="#" * 12)),
                        cell=translate(i, 0),
                    ),
                    product_description=TextLabel(
                        value=fake.text(max_nb_chars=40)
                        .replace(" ", "  ")[:28]
                        .upper(),
                        cell=translate(i, 1),
                    ),
                    site=NumberLabel(
                        value=fake.random_int(10, 999), cell=translate(i, 2)
                    ),
                    zone=TextLabel(
                        value=fake.bothify(text="#" * 7), cell=translate(i, 3)
                    ),
                    date_range=TextLabel(
                        value=fmt.american_date(
                            fake.date_between(start_date=DocType.START_DATE)
                        )
                        + " "
                        + fmt.american_date(
                            fake.date_between(start_date=DocType.START_DATE)
                        ),
                        cell=translate(i, 4),
                    ),
                    scan_dollars=NumberLabel(
                        value=fake.random_number(digits=3) / 100.0, cell=translate(i, 5)
                    ),
                    units=NumberLabel(
                        value=fake.random_number(digits=1), cell=translate(i, 6)
                    ),
                    amount_dollars=NumberLabel(
                        value=fake.random_number(digits=3) / 100.0, cell=translate(i, 7)
                    ),
                    scan_deal=NumberLabel(
                        value=opts.maybe_empty(fake.random_number(digits=7)),
                        cell=translate(i, 8),
                    ),
                )
                for i in range(
                    opts.num_rows
                    if opts.num_rows is not None
                    else fake.random_int(50, 200)
                )
            ],
        )

    @classmethod
    def generate_wb(cls, opts: GenOpts):
        wb = load_workbook(DocType.get_template_path("d1.xlsx"))
        ws = wb[wb.sheetnames[0]]

        record = cls.generate_record(opts)
        write_cells(ws, record)

        first_table_row = cls.FIRST_TABLE_ROW
        first_table_column = cls.FIRST_TABLE_COLUMN

        n_cols = len(dict(record.rows[0]))

        # Copy the first row's style to each row
        for n, row in enumerate(record.rows):
            if n == 0:
                continue

            for x in range(n_cols):
                first_row_cell = ws.cell(first_table_row, first_table_column + x)
                cell = ws.cell(first_table_row + n, first_table_column + x)
                if first_row_cell.has_style:
                    # https://stackoverflow.com/questions/23332259/copy-cell-style-openpyxl
                    cell._style = copy(first_row_cell._style)

        # Adjust column widths to be very narrow
        for col_num in range(first_table_column, first_table_column + n_cols):
            max_width = 0
            for row_num in range(
                first_table_row - 2, first_table_row + len(record.rows) + 2
            ):
                cell = ws.cell(row_num, col_num)
                if cell.value:
                    width, _ = text_dims(cell.font.name, cell.font.size, cell.value)
                    max_width = max(width, max_width)

            # https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size
            # https://stackoverflow.com/questions/60264964/what-is-the-unit-used-in-openpyxl-for-column-width
            #   (px-5)/6 = units in this sheet. Every sheet may be different based on the default font.
            # Excel also subtracts about 0.83 units from the width automatically
            ws.column_dimensions[get_column_letter(col_num)].width = (
                (max_width + 3 - 5) / 6
            ) + 0.83

        return record, wb

    @classmethod
    @validate_arguments
    def generate_doc(
        cls, dstdir: pathlib.PosixPath, opts: GenOpts, config: Config, i: int
    ):
        log = config.logger("d1-%d" % (i))
        log.debug("Starting!")

        record, wb = cls.generate_wb(opts)

        pdf_path = dstdir.joinpath("d1-%d.pdf" % (i + 1))
        write_wb_to_pdf_path(wb, pathlib.Path(pdf_path))

        log.debug("Generated workbook")

        if opts.build_training_data:
            annotate_locations(wb, record)

            if opts.remove_multipage_rows:
                remove_multipage_rows(clone_wb(wb), record)
                write_wb_to_pdf_path(wb, pathlib.Path(pdf_path))

            with open(dstdir.joinpath("d1-%d.json" % (i + 1)), "w") as f:
                f.write(record.json(indent=2))

        log.debug("Done!")
