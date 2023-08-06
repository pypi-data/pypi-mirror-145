from datetime import date, datetime
from openpyxl import Workbook, load_workbook
import pathlib
from pydantic import BaseModel, validate_arguments
import random
from typing import Any, List

from .. import fmt
from ..types import (
    Cell,
    Config,
    GenOpts,
    NumberLabel,
    ScalarLabel,
    TextLabel,
    TimestampLabel,
    traverse,
)
from ..xl import (
    annotate_locations,
    copy_style_range,
    clone_wb,
    write_wb_to_pdf_path,
    write_cells,
    remove_multipage_rows,
)
from .doc_type import DocType


class D2Row(BaseModel):
    claimant_name: TextLabel
    claim_number: TextLabel
    loss_date: TimestampLabel
    loss_state: TextLabel
    receipt_date: TimestampLabel
    div_number: NumberLabel
    status: TextLabel
    closed_date: TimestampLabel
    adjuster_name: TextLabel
    manager_name: TextLabel
    loss_description: TextLabel
    loss_paid: NumberLabel


class D2Record(BaseModel):
    policy_number: TextLabel
    company: TextLabel
    date_range: TextLabel
    report_time: TimestampLabel

    rows: List[D2Row]


class D2(DocType):
    DOC_NUMBER = 2
    Record = D2Record

    FIRST_TABLE_ROW = 11
    FIRST_TABLE_COLUMN = 1
    ROW_HEIGHT = 4

    @classmethod
    @validate_arguments
    def generate_record(cls, opts: GenOpts) -> D2Record:
        fake = DocType.faker()

        def translate(row_index, sub_row_index, col_index, **kwargs):
            return Cell(
                row=cls.FIRST_TABLE_ROW
                + cls.ROW_HEIGHT * row_index
                + 1
                + sub_row_index,
                column=cls.FIRST_TABLE_COLUMN + col_index,
                **kwargs
            )

        date_range_start = fake.date_between(start_date=DocType.START_DATE)
        date_range_end = fake.date_between(start_date=date_range_start)

        num_rows = (
            opts.num_rows if opts.num_rows is not None else fake.random_int(50, 80)
        )
        loss_dates = [
            fake.date_between(start_date=date_range_start, end_date=date_range_end)
            for x in range(num_rows)
        ]
        loss_dates.sort()

        receipt_dates = [
            fake.date_between(start_date=ld, end_date=date_range_end)
            for ld in loss_dates
        ]
        closed = [
            random.choice([True, False]) if opts.optional_vals else True
            for _ in receipt_dates
        ]
        closed_dates = [
            fake.date_between(start_date=rd, end_date=date_range_end) if cl else None
            for cl, rd in zip(closed, receipt_dates)
        ]

        return D2Record(
            policy_number=TextLabel(
                value=fake.bothify(text="##########-###-###"),
                cell=Cell(row=5, column=2),
            ),
            company=TextLabel(value=fake.company(), cell=Cell(row=6, column=1)),
            date_range=TextLabel(
                value=fmt.american_date(date_range_start)
                + " - "
                + fmt.american_date(date_range_end),
                cell=Cell(row=6, column=2),
            ),
            report_time=TimestampLabel(
                value=datetime.combine(
                    fake.date_between(start_date=date_range_end), datetime.min.time()
                ),
                cell=Cell(row=5, column=6),
            ),
            rows=[
                D2Row(
                    claimant_name=TextLabel(value=fake.name(), cell=translate(i, 0, 0)),
                    claim_number=TextLabel(
                        value=fake.bothify(text="###-######-###"),
                        cell=translate(i, 1, 0),
                    ),
                    loss_date=TimestampLabel.from_date(
                        loss_dates[i], cell=translate(i, 2, 0)
                    ),
                    loss_state=TextLabel(
                        value=fake.state_abbr(), cell=translate(i, 1, 1)
                    ),
                    receipt_date=TimestampLabel.from_date(
                        receipt_dates[i], cell=translate(i, 2, 1)
                    ),
                    div_number=NumberLabel(
                        value=fake.random_number(digits=3), cell=translate(i, 0, 2)
                    ),
                    status=TextLabel(
                        value="Closed" if closed[i] else "Open", cell=translate(i, 1, 2)
                    ),
                    closed_date=TimestampLabel.from_date(
                        closed_dates[i], cell=translate(i, 2, 2)
                    ),
                    adjuster_name=TextLabel(value=fake.name(), cell=translate(i, 0, 3)),
                    manager_name=TextLabel(value=fake.name(), cell=translate(i, 1, 3)),
                    loss_description=TextLabel(
                        value=fake.sentence(10), cell=translate(i, 0, 4)
                    ),
                    loss_paid=NumberLabel(
                        value=fake.random_number(digits=fake.random_int(min=0, max=5))
                        / 100.0,
                        cell=translate(i, 2, 5),
                    ),
                )
                for i in range(num_rows)
            ],
        )

    @classmethod
    def generate_wb(cls, opts: GenOpts):
        wb = load_workbook(DocType.get_template_path("d2.xlsx"))
        ws = wb[wb.sheetnames[0]]

        record = cls.generate_record(opts)
        write_cells(ws, record)

        first_table_row = cls.FIRST_TABLE_ROW
        first_table_column = cls.FIRST_TABLE_COLUMN

        n_cols = len(dict(record.rows[0]))

        # Copy the first row's style to each row
        for n, row in enumerate(record.rows):
            if n == 0:
                continue

            start_row = first_table_row + n * cls.ROW_HEIGHT
            ws.merge_cells(
                start_row=start_row + 1,
                start_column=5,
                end_row=start_row + 3,
                end_column=5,
            )
            copy_style_range(
                ws,
                first_table_row,
                first_table_column,
                cls.ROW_HEIGHT,
                6,
                start_row,
                first_table_column,
            )

        return record, wb

    @classmethod
    @validate_arguments
    def generate_doc(
        cls, dstdir: pathlib.PosixPath, opts: GenOpts, config: Config, i: int
    ):
        log = config.logger("d2-%d" % (i))

        record, wb = cls.generate_wb(opts)

        pdf_path = dstdir.joinpath("d2-%d.pdf" % (i + 1))
        write_wb_to_pdf_path(wb, pathlib.Path(pdf_path))

        if opts.build_training_data:
            annotate_locations(clone_wb(wb), record)

            if opts.remove_multipage_rows:
                remove_multipage_rows(wb, record)
                write_wb_to_pdf_path(wb, pathlib.Path(pdf_path))

            with open(dstdir.joinpath("d2-%d.json" % (i + 1)), "w") as f:
                f.write(record.json(indent=2))
