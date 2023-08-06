import os
import struct
import openpyxl
import numpy as np
from pathlib import Path
from typing import Dict, List
from skimage import io
import matplotlib.pyplot as plt
from scipy import signal as sgl
import test_material_brain as mb


# FIGSIZE = (14, 7)
# LC_LINE_COLOR = 'r'
# FIRST_PEAK_LINE_COLOR = 'g'

__all__ = ['Scratch']

class Scratch:
    """
    Represent a scratch measurment.

    Attributes
    -----------------------------------------------
        filename : Measurement file name

        Name : Measurement name

        panorama : Scratch panorama image   [Col:Row:RGB]

        signals : Measured signals (dictionary)
            - Ft  Frictional force          [mN]
            - Ae  Acoustic emission         [%]
            - Pd  Penetration depth         [nm]
            - FnC Command normal force      [mN]
            - Pf  Surface profile           [nm]
            - Rd  Residual depth            [nm]
            - Fn  Measured normal force     [mN]
            - mu  Friction coefficient      [-]

        fs : Sampling frequency             [Hz]

        pos : vector of position            [mm]

        t : vector of time                  [s]

        lcs : Critical loads (dictionary)
            - LC1  Crack                    [mN]
            - LC2  Cohesive spallation      [mN]
            - LC3  Interfacial spallation   [mN]
            - LC4  Delamination             [mN]
            - LC5  Not defined              [mN]

        signalFullNames : Full names of the signals (dictionary)

        signalUnits : Units of the signals (dictionary)

        signalColors : Colors of the signals (dictionary)

        lcsNames : Names of critical loads (dictionary)
    """
    LC1 = "LC1"
    LC2 = "LC2"
    LC3 = "LC3"
    LC4 = "LC4"
    
    FT = 'Ft'   # Frictional force       [mN]
    AE = 'Ae'   # Acoustic emission      [%]
    PD = 'Pd'   # Penetration depth      [nm]
    FN = 'Fn'   # Normal force (measured)[mN]
    PF = 'Pf'   # Profile                [nm]
    RD = 'Rd'   # Residual depth         [nm]
    FNC = 'FnC' # Commanded normal force [mN]
    MU = 'mu'   # Friction coefficient   [-]
    LABEL = 'label'


    def __init__(self, filename, panorama, signals, pos, t, lcs):
        self.filename = os.path.basename(filename)
        self.name = Path(filename).stem
        self.signals = signals
        self.panorama = panorama
        self.pos = pos
        self.t = t
        self.lcs = lcs
        self.fs = mb.ComputeFs(self.t)

        # Constants
        self.signalFullNames = {
            self.FT: 'Frictional force',
            self.AE: 'Acoustic emission',
            self.PD: 'Penetration depth',
            self.FNC: 'Command normal force',
            self.PF: 'Surface profile',
            self.RD: 'Residual depth',
            self.FN: 'Measured normal force',
            self.MU: 'Friction coefficient'}
        self.signalUnits = {
            self.FT: 'mN',
            self.AE: '%',
            self.PD: 'nm',
            self.FNC: 'mN',
            self.PF: 'nm',
            self.RD: 'nm',
            self.FN: 'mN',
            self.MU: '-'}
        self.signalColors = {
            self.FT: 'purple',
            self.AE: 'b',
            self.PD: 'g',
            self.FNC: 'gray',
            self.PF: 'olive',
            self.RD: 'lightgreen',
            self.FN: 'brown',
            self.MU: 'r'}
        self.lcsNames = {
            self.LC1: 'Crack',
            self.LC2: 'Cohesive spallation',
            self.LC3: 'Interfacial spallation',
            self.LC4: 'Delamination',
            'LC5': '-'}
   
    @classmethod
    def retrieve_data(cls, filePath, getPanorama = True, imagePath = None, imageExtension = ".bmp", printInfos = False)->'Scratch':
        """
        ### Function to retrieve scratch signals and critical loads from a .xlsx file.
            This function can also retrieve a panorama image.
        
        #### Inputs:
        - filePath          File path (must be an .xlsx file)
        - getPanorama       If true, retrieve panorama image
        - imagePath         Image path
        - imageExtension    Image extension
        - printInfos        If true, print some informations about this scratch

        #### Outputs:
        - scratch           An object scratch that contains all the needed data and methods
        """

        # Constants
        lc_first_row = 2
        lc_last_row = 6
        signal_first_row = 9

        coef_mN = 1000
        coef_mm = 1000
        coef_nm = 1000000
        coef_Ae = 100

        # panorama .bmp file info
        pixel_per_meter_byte_start = 38
        pixel_per_meter_byte_end = 42
        
        # Vectors
        t   = [] # Time [s]
        pos = [] # Position [mm]
        
        # Critical loads
        lcs = {} # [mN]

        # Signals    
        signals = {Scratch.FT  : [], # Frictional force       [mN]
                Scratch.AE  : [], # Acoustic emission      [%]
                Scratch.PD  : [], # Penetration depth      [nm]
                Scratch.FNC : [], # Command normal force           [mN]
                Scratch.PF  : [], # Profile                [nm]
                Scratch.RD  : [], # Residual depth         [nm]
                Scratch.FN  : [], # Measured normal force  [mN]
                Scratch.MU  : []} # Friction coefficient   [-]
        
        workbook = openpyxl.load_workbook(filename = filePath, read_only = True, data_only = True)
        sheet = workbook.active
        
        # Retrieve critical loads
        for idx, row in enumerate(sheet.iter_rows(min_row = lc_first_row, max_row = lc_last_row, values_only = True), start = 1):
            if row[0] != 'nan':
                lcs[f'{idx}'] = row[0]

        # Retrieve signals
        for row in sheet.iter_rows(min_row = signal_first_row, values_only = True):
            pos.append(row[0] * coef_mm)
            signals[Scratch.FT].append(row[1] * coef_mN)
            signals[Scratch.AE].append(row[2] * coef_Ae)
            signals[Scratch.PD].append(row[3] * coef_nm)
            signals[Scratch.FNC].append(row[4] * coef_mN)
            signals[Scratch.PF].append(row[5] * coef_nm)
            signals[Scratch.RD].append(row[6] * coef_nm)
            t.append(row[7])
            signals[Scratch.FN].append(row[8] * coef_mN)
            signals[Scratch.MU].append(row[11])

        # Convert each vector in numpy array
        t = np.array(t)
        pos = np.array(pos)

        # Convert each signal in numpy array
        for name, signal in signals.items():
            signals[name] = np.array(signal)

        # Retrieve Panorama (optional)
        if getPanorama:
            if imagePath == None:
                imagePath = os.path.splitext(filePath)[0] + imageExtension

            panorama = []
            try:
                panorama = io.imread(imagePath) 
                if imageExtension == '.bmp':
                    pixelPerMeter = int(struct.unpack('I', open(imagePath, "rb").read()[pixel_per_meter_byte_start:pixel_per_meter_byte_end])[0])
                    dataLength = (pos[-1] - pos[0])
                    panorama = mb.cropImage(panorama, dataLength, pixelPerMeter)

            except FileNotFoundError:
                print(f"Could not find an image with this path: {imagePath}\n")

            except:
                print ("An error occured with the panorama.\n")

            if len(panorama) == 0:
                panorama = None
            
        else:
            panorama = None

        # Create scratch object
        scratch = cls(filePath, panorama, signals, pos, t, lcs)
        
        # Display scratch informations (optional)
        if printInfos:
            print(scratch)
            
        return scratch
    
    def __str__(self):
        return f"------------------------------------------------------------------------------------------------------------\n" \
            f" {self.filename}\n" \
            f"------------------------------------------------------------------------------------------------------------\n" \
            f" Nb. sample:\t{self.length()}\n" \
            f" Duration:\t{self.duration():.3f} \ts\n" \
            f" Distance:\t{self.distance():.3f} \tmm\n" \
            f" Fs:\t\t{self.fs:.3f} \tHz\n\n" \
            f" Panorama\n" \
            f" - Panorama:\t{'Yes' if self.has_panorama() else 'No'}\n\n" \
            f" Critical loads\n" \
            f" - LC1:\t\t{'Yes' if self.has_lc(1) else 'No'}\t({self.lcsNames['LC1']})\n" \
            f" - LC2:\t\t{'Yes' if self.has_lc(2) else 'No'}\t({self.lcsNames['LC2']})\n" \
            f" - LC3:\t\t{'Yes' if self.has_lc(3) else 'No'}\t({self.lcsNames['LC3']})\n" \
            f" - LC4:\t\t{'Yes' if self.has_lc(4) else 'No'}\t({self.lcsNames['LC4']})\n" \
            f" - LC5:\t\t{'Yes' if self.has_lc(5) else 'No'}\t(Not defined)\n\n" \
            f" Signals\n" \
            f" - Ft:\t\t{'Yes' if self.has_signal('Ft') else 'No'}\t({self.signalFullNames['Ft']})\n" \
            f" - Ae:\t\t{'Yes' if self.has_signal('Ae') else 'No'}\t({self.signalFullNames['Ae']})\n" \
            f" - Pd:\t\t{'Yes' if self.has_signal('Pd') else 'No'}\t({self.signalFullNames['Pd']})\n" \
            f" - FnC:\t\t{'Yes' if self.has_signal('FnC') else 'No'}\t({self.signalFullNames['FnC']})\n" \
            f" - Pf:\t\t{'Yes' if self.has_signal('Pf') else 'No'}\t({self.signalFullNames['Pf']})\n" \
            f" - Rd:\t\t{'Yes' if self.has_signal('Rd') else 'No'}\t({self.signalFullNames['Rd']})\n" \
            f" - Fn:\t\t{'Yes' if self.has_signal('Fn') else 'No'}\t({self.signalFullNames['Fn']})\n" \
            f" - mu:\t\t{'Yes' if self.has_signal('mu') else 'No'}\t({self.signalFullNames['mu']})\n" \
            f"------------------------------------------------------------------------------------------------------------\n"

    def length(self):
        """Gets number of samples"""
        return len(self.t)

    def duration(self):
        """Gets scratch duration [s]"""
        return self.t[-1] - self.t[0]

    def distance(self):
        """Gets scratch lengthwise distance [mm]"""
        return self.pos[-1] - self.pos[0]

    def loading_rate(self):
        """Gets loading rate [mN/s]"""
        return (self.signals[self.FN][-1]-self.signals[self.FN][0])/self.duration()
    
    def speed(self):
        """Gets speed [mm/s]"""
        return self.distance()/self.duration()

    def get_only_neccesary_signal_parts(self, signal_names:List[str], lc_search:int, lc_positions:Dict[int, int], lc_index_for_label_signal:int = -1) -> Dict[str, List[float]]:
        """ Removes unnecessary parts of the signals depending on the critical load sought 
        and the critical loads already found and returns it.

        Args:
            signal_names (List[str]): The names of the signals
            lc_search (int): The LC to search
            lc_positions (Dict[int, int]): LC already found. the key is the lc number and the value is the index.
            lc_index_for_label_signal (int, optional): The index of the LC to search. It's to add the labelization signal for training dataset. 
            -1 = no labelization. Defaults to -1.

        Returns:
            Dict[str, List[float]]): A dictionnary. the key is the name of the signal. the value is a list of float corresponding to the signal.
        """
        dict_signals = {}
        start_index = 0
        end_index = self.length()

        # Get the lc's that will enable the signals to be cut according to the critical load sought
        lc_limits_for_lc_search = Scratch.get_lcs_to_extract(lc_search)
        start_lc = lc_limits_for_lc_search[0]
        end_lc = lc_limits_for_lc_search[1]

        # Find the start index for the signals cutting
        if start_lc > 0:
            if start_lc in lc_positions:
                start_index = lc_positions[start_lc]
            else:
                LcNotDefinedError(start_lc)

        # Find the end index for the signals cutting
        if end_lc > 0:
            if end_lc in lc_positions:
                end_index = lc_positions[end_lc]
            else:
                LcNotDefinedError(end_lc)

        # Cut the signals and put it in the dictionnary
        for signal_name in signal_names:
            signal = self.get_signal(signal_name)
            dict_signals[signal_name] = signal[start_index:end_index]

        # Create labelization signals if requested
        if lc_index_for_label_signal >= 0:
            labels = self.get_labelization(lc_index_for_label_signal)
            dict_signals[Scratch.LABEL] = labels[start_index:end_index]
            
        return dict_signals

    def get_labelization(self, lc_index:int) -> List[int]:
        """Creates a standard labeling for the lc index. 
        The value 0 is set before the critical load and 1 after the critical load. 

        Args:
            lc_index (int): The index of the critical load

        Returns:
            List[int]: The labelization
        """
        signal_length = self.length()
        labels = np.zeros(signal_length, dtype=int)
        for i in range(signal_length - 1):
            if i >= lc_index:
                labels[i] = 1
        return labels

    def get_signal(self, signal_name:str):
        """_summary_

        Args:
            signal_name (str): _description_

        Raises:
            SignalNotDefinedError: _description_

        Returns:
            _type_: _description_
        """
        if not self.has_signal(signal_name):
            raise SignalNotDefinedError(self, signal_name)
        return self.signals[signal_name]

    def lcs_index(self):
        """Gets critical loads's index"""
        index = []

        for lc in self.lcs.values():
                index.append(mb.findNearestValue(self.signals[self.FN], lc))

        return index

    def lc_index(self, lc):
        """
        ### Gets the index of a particular critical load 

        #### Inputs:
        - lc        Number of the critical load (ex. 1 for LC1)

        #### Outputs:
        - index     Index of the critical load
        """
        
        if not self.has_lc(lc):
            raise LcNotDefinedError(self, lc)

        return mb.findNearestValue(self.signals[self.FN], self.lcs[str(lc)])

    def lc_index_panorama(self, lc):
        """
        ### Gets the index of a particular critical load on panorama

        #### Inputs:
        - lc        Number of the critical load (ex. 1 for LC1)

        #### Outputs:
        - index     Index of the critical load
        """

        if not self.has_lc(lc):
            raise LcNotDefinedError(self, lc)
        if not self.has_panorama():
            raise PanoramaNotDefinedError(self)
        
        index = self.lc_index(lc)
        return int(index / self.length() * self.panorama.shape[1])

    def has_panorama(self):
        """
        ### Check if a panorama has been recorded for this scratch

        #### Outputs:
        - bool      True if the panorama has been recorded
        """
        return not self.panorama is None

    def has_lc(self, lc):
        """
        ### Check if a particular critical load has been defined for this scratch
        
        #### Inputs:
        - lc        Number of the critical load (ex. 1 for LC1)

        #### Outputs:
        - bool      True if the lc has been defined
        """
        return str(lc) in self.lcs

    def has_lcs(self, lcs):
        """
        ### Check if some particular critical loads have been defined for this scratch
        
        #### Inputs:
        - lcs       Numbers of the critical loads to be checked (ex. [1,3] to check LC1 and LC3)

        #### Outputs:
        - bool      True if the lcs have been defined
        """
        for lc in lcs:
            if not self.has_lc(lc):
                return False
        return True

    def has_signal(self, signal):
        """
        ### Check if a particular signal has been recorded for this scratch
        
        #### Inputs:
        - signal    Name of the signal (ex. 'Ae')

        #### Outputs:
        - bool      True if the signal has been recorded
        """
        return self.signals[signal].any()
    
    def has_signals(self, signals = [FT, AE, PD, FN, PF, RD, FNC, MU]):
        """
        ### Check particular signals has been recorded for this scratch
        
        #### Inputs:
        - signals   Name of the signals, can be : 'Ft', 'Ae', 'Pd', 'FnC', 'Pf', 'Rd', 'Fn' or 'mu' (ex. [Ft, Pd] to check Ft and Pd)

        #### Outputs:
        - bool      True if the signal has been recorded
        """
        for signal in signals:
            if not self.has_signal(signal):
                return False
        return True

    def has_data(self, signals = [FT, AE, PD, FN, PF, RD, FNC, MU], lcs = [1, 2, 3, 4], check_panorama = True):
        """
        ### Function to check scratch content

        #### Inputs:
        - signals           Name of the signals to be checked, can be : 'Ft', 'Ae', 'Pd', 'FnC', 'Pf', 'Rd', 'Fn' or 'mu' (ex. [Ft, Pd] to check Ft and Pd)
        - lcs               Numbers of the critical loads to be checked (ex. [1,3] to check LC1 and LC3)
        - check_panorama    Boolean, if true check panorama image
        """

        is_missing = False

        # Check signals
        if not self.has_signals(signals):
            is_missing = True

        # Check critical loads
        if not self.has_lcs(lcs):
            is_missing = True

        # Check panorama image
        if check_panorama:
            if not self.has_panorama():
                is_missing = True

        return not is_missing

    @staticmethod
    def get_lcs_to_extract(lc_search:int) -> List[int]:
        """_summary_

        Parameters
        ----------
        lc_search : int
            Lc search for which the lcs to extract are defined

        Returns
        -------
        List[int]
            Lcs to extract of length 2 from start to end 
            0 : from start of data
            -1 :to end of data 
            1 - 4 : from/to lc_index
        """
        if lc_search == 1:
            return [0,2]
        elif lc_search == 2:
            return [0,-1]
        elif lc_search == 3:
            return [2,4]
        elif lc_search == 4:
            return [2,-1]

    """
    Display functions
    """  

    def find_first_peak(self, signal_name, prominence, plot = False, x_axis = 'position', x_min = None, x_max = None, show_lcs = False, show_lc_nb = None, figsize = mb.FIGSIZE):
        """
        ### Find first peak in a signal
        
        #### Inputs:
        - signal_name   Name of the signal to find first peak, can be : 'Ft', 'Ae', 'Pd', 'FnC', 'Pf', 'Rd', 'Fn' or 'mu'
        - prominence    Required prominence of peaks
        - plot          Boolean, if true plot the signal with his maxima
        - x_axis        Name of the x-axis, can be : 'time', 'position' or 'none'
        - x_min         Index of the minimum value for the x-axis
        - x_max         Index of the maximum vlaue for the x-axis
        - show_lcs      Boolean, if true display all critical loads
        - show_lc_nb    Number of the critical load to display on graphic
        - figsize       Size of the figure


        #### Outputs:
        - index         Index of the first peak found
        """

        # Find maxima and the index of the first peak
        signal = self.signals[signal_name][x_min:x_max]
        normalized_signal = mb.normalize(signal)
        peaks, _ = sgl.find_peaks(normalized_signal, prominence = prominence)

        if len(peaks) > 0:
            index = peaks[0]
        else:
            index = -1

        if plot:
            fig, ax = self.plot_signal(signal_name, x_axis = x_axis, x_min = x_min, x_max = x_max,show = False, show_lcs = show_lcs, show_lc_nb = show_lc_nb, figsize = figsize)

            if x_axis == 'time':
                x = self.t
            elif x_axis == 'position':
                x = self.pos
            else:
                x = None

            # Plot maxima and first peak line
            if x is None:
                ax.scatter(peaks, signal[peaks], color = 'r', s = 10, marker = 'D', label = 'Maxima')
                ax.axvline(index, c = mb.FIRST_PEAK_LINE_COLOR, label = 'First peak')
            else:
                ax.scatter(x[peaks], signal[peaks], color = 'r', s = 10, marker = 'D', label = 'Maxima')
                ax.axvline(x[index], c = mb.FIRST_PEAK_LINE_COLOR, label = 'First peak')

            # Show figure
            plt.legend(loc='best')
            plt.show()

        return index

    def plot_signal(self, signal_name, func = None, x_axis = 'position', x_min = None, x_max = None, show = True, display_lcs = None, figsize = mb.FIGSIZE):
        """
        ### Dsiplay a scratch signal
        
        #### Inputs:
        - signal_name   Name of the signal to be displayed, can be : 'Ft', 'Ae', 'Pd', 'FnC', 'Pf', 'Rd', 'Fn' or 'mu'
        - func          Function to apply to the selected signal (ex. func = lambda signal : process(signal, param1, param2))
        - x_axis        Name of the x-axis, can be : 'time', 'position' or 'none'
        - x_min         Index of the minimum value for the x-axis
        - x_max         Index of the maximum vlaue for the x-axis
        - show          Boolean, if true display the figure
        - display_lcs   Numbers of the critical loads to display (ex. display_lcs = [1,3] to dislpay LC1 and LC3)
        - figsize       Size of the figure

        #### Outputs:
        - fig           figure
        - ax            axis
        """

        if not func is None:
            signal = func(self.signals[signal_name][x_min:x_max])
        else:
            signal = self.signals[signal_name][x_min:x_max]

        fig, ax = plt.subplots(figsize = figsize)
        ax.grid(b = True, which = 'major', axis = 'both')

        if x_axis == 'time':
            ax.plot(self.t[x_min:x_max], signal, label = f'{signal_name}')
            ax.set_xlabel('time [s]')
        elif x_axis == 'position':
            ax.plot(self.pos[x_min:x_max], signal, label = f'{signal_name}')
            ax.set_xlabel('position [mm]')
        else:
            ax.plot(signal, label = f'{signal_name}')
            ax.set_xlabel('samples [-]')

        ax.set_ylabel(f'{signal_name} [{self.signalUnits[signal_name]}]')
        ax.margins(x=0, y=0)


        # Plot critical loads
        if not display_lcs is None:
            mb.plot_lcs(self, ax, display_lcs, x_axis, color = mb.LC_LINE_COLOR)

        if show:
            plt.legend(loc = 'best')
            plt.show()

        return fig, ax

    def plot_signals(self, signals = [FT, AE, PD, RD, FN, MU], func = None, x_axis = 'position', x_min = None, x_max = None, y_axis = True, show = True, display_lcs = None, figsize = mb.FIGSIZE):
        """
        ### Display all the signals of a scratch except FnC and Pf
        
        #### Inputs:
        - signals       Name of the signals to be displayed, can be : 'Ft', 'Ae', 'Pd', 'FnC', 'Pf', 'Rd', 'Fn' or 'mu'
        - func          Function to apply to the selected signal (ex. func = lambda signal : process(signal, param1, param2))
        - x_axis        Name of the x-axis, can be : 'time', 'position' or 'none'
        - x_min         Index of the minimum value for the x-axis
        - x_max         Index of the maximum vlaue for the x-axis
        - y_axis        Boolean, if true display on y-axis labels and ticks values
        - show          Boolean, if true display the figure
        - display_lcs   Numbers of the critical loads to display (ex. display_lcs = [1,3] to dislpay LC1 and LC3)
        - figsize       Size of the figure

        #### Outputs:
        - fig           figure
        - ax            axis
        - ps            
        """

        # Figure settings
        fig, ax1 = plt.subplots(figsize = figsize)
        fig.suptitle(self.filename, fontsize=16)
        ax1.grid(b = True, which = 'major', axis = 'both')

        # Variables
        i = 0
        ps = []
        axs = [ax1]
        
        # For each signal
        for item in self.signals.items():

            signal_name = item[0]
            signal = item[1]

            if self.has_signal(signal_name) and signal_name in signals:

                # Applies signal processing if defined
                if not func is None:
                    signal = func(signal[x_min:x_max])
                else:
                    signal = signal[x_min:x_max]

                # Adds y-axis on the right of the graphic
                if i > 0:
                    axs.append(ax1.twinx())
                    if y_axis:
                        axs[i].spines["right"].set_position(("axes", 1 + 0.08 * (i-1)))
               
                # Plot signal
                if x_axis == 'time':
                    p, = axs[i].plot(self.t[x_min:x_max], signal, self.signalColors[signal_name], label = f'{signal_name}')
                    axs[i].set_xlabel('time [s]')
                elif x_axis == 'position':
                    p, = axs[i].plot(self.pos[x_min:x_max], signal, self.signalColors[signal_name], label = f'{signal_name}')
                    axs[i].set_xlabel('position [mm]')
                else:
                    p, = axs[i].plot(signal, self.signalColors[signal_name], label = f'{signal_name}')
                    axs[i].set_xlabel('samples [-]')

                if signal_name == 'Pf' or signal_name == 'Pd' or signal_name == 'Rd':
                    axs[i].invert_yaxis()
                
                # Y-axis settings
                if y_axis: 
                    axs[i].yaxis.label.set_color(p.get_color())
                    # axs[i].tick_params(axis='y', colors=p.get_color())
                    axs[i].set_ylabel(f'{signal_name} [{self.signalUnits[signal_name]}]')
                else:
                    axs[i].set_yticks([])

                # Axis settings
                ps.append(p)
                axs[i].margins(x=0, y=0)

                i += 1 # Increment counter only if a signal is plotted

        # Plot critical loads
        if not display_lcs is None:
            mb.plot_lcs(self, ax1, display_lcs, x_axis, color = mb.LC_LINE_COLOR)

        # Show figure
        if show:
            ax1.legend(handles = ps, loc = 'best')
            plt.show()

        return fig, ax1, ps

    def display_panorama(self, downscale_ratio = 10, title = False, show_lcs = True, show = True, borne=[0,-1,0,-1]):
        # if downscaleRatio > 1 :
        #     image = transform.downscale_local_mean(np.array(self.panorama), (downscaleRatio, downscaleRatio, 1))/255
        # else : 
        #     image = self.panorama/255

        image = mb.downscale_panorama(self.panorama, downscale_ratio)/255
        
        fig, ax = plt.subplots(figsize = (18, 9), clear = True)
        ax.set_xmargin(0)
        ax.set_yticks([])
        ax.set_xticks([])
        plt.box(on = None)
        if title : plt.title(self.filename , fontsize=15, color = 'white')
        io.imshow(image[borne[0]:borne[1],borne[2]:borne[3]])

        if show_lcs:
            mb.plot_lc_panorama(self, ax, len_xAxis=np.shape(image)[1], start_ind=borne[2], end_ind=borne[3])

        if show: plt.show()

        return fig, ax, image



class MaterialBrainError(Exception):
    """
    Material Brain base class for exceptions
    """
    def __init__(self, message):
        self.message = message

    def __str__(self):
        return self.message

class LcNotDefinedError(MaterialBrainError):
    def __init__(self, scratch, lc):

        self.scratch = scratch
        self.lc = lc

        if lc > 0 and lc < 6 :
            self.message = f'Requested critical load (n° {self.lc}) is not defined for this scratch : {self.scratch.filename}'
        else:
            self.message = f'Requested critical load (n° {self.lc}) is not in 1-5 range.'

        super().__init__(self.message)

class SignalNotDefinedError(MaterialBrainError):
    def __init__(self, scratch, signal):

        self.scratch = scratch
        self.signal = signal

        self.message = f'Requested signal ({self.signal}) is not defined for this scratch : {self.scratch.filename}'

        super().__init__(self.message)

class PanoramaNotDefinedError(MaterialBrainError):
    def __init__(self, scratch):

        self.scratch = scratch

        self.message = f'Requested panorama image is not defined for this scratch : {self.scratch.filename}'

        super().__init__(self.message)
