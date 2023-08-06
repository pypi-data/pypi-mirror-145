import ast
import json
import csv
import openpyxl
from typing import Callable, Any


def decorator(func: Callable) -> Callable:
    def check_repeat(*args, **kwargs: Any) -> Any:
        """
        Takes main function, repeats it if necessary.
        Cuts Errors, then stops.

        :return: finally returns None.
        """
        while True:
            try:
                result = func(*args, **kwargs)
                while True:
                    answer = input("Program has finished it's work. Do you want to repeat? (Y/N): ").upper()
                    if answer == 'Y':
                        break
                    elif answer != 'N':
                        continue
                    else:
                        return result

            except KeyError:
                print("This is not a path to a file with it's name.\n"
                      "The name should end with .txt, .json, .csv, .xlsx.")
                break
            except FileNotFoundError:
                print("Such path to the file does not exist.")
                break
            except SyntaxError:
                print("Your .txt file just imitates list with dicts inside, check it again.")
                break
            except PermissionError:
                break
            except Exception:
                print('Something went wrong. Check everything once again.')
                break

    return check_repeat


def from_txt(path: str) -> list or None:
    """
    .txt file should have
    ONLY LIST WITH DICTS in str inside.

    :param path: path to the source .txt.
    :return: list with dicts or None if Error.
    """
    with open(path, 'r', encoding='utf-8') as file:
        inhalt = file.read().strip()
        if inhalt[:2] != '[{' or inhalt[-2:] != '}]':
            print('Your source .txt file must consist of list with dicts inside.')
            return None
        return ast.literal_eval(inhalt)


def from_json(path: str) -> list:
    with open(path, 'r', encoding='utf-8') as file:
        return json.loads(file.read())


def from_csv(path: str) -> list:
    with open(path, 'r', encoding='utf-8') as file:
        inner_lst = []
        reader = csv.reader(file, delimiter=',')
        for i, row in enumerate(reader):
            if i == 0:
                keys = row
                continue
            inner_lst.append(dict(zip(keys, row)))
        return inner_lst


def from_xlsx(path: str) -> list:
    wb = openpyxl.load_workbook(path)
    name_sheet = input('Type the name of the sheet to write the data from: ')
    try:
        sheet = wb[name_sheet]
        inner_lst = []
        keys = tuple()
        for i, row in enumerate(sheet.values):
            if i == 0:
                keys = row
                continue
            inner_lst.append(dict(zip(keys, row)))
        return inner_lst
    except KeyError:
        print(f'Worksheet {name_sheet} does not exist.')
        return from_xlsx(path)


def to_txt(py_list: list, path: str) -> None:
    new_path = path[:path.rfind('.') + 1] + 'txt'
    with open(new_path, 'w', encoding='utf-8') as file:
        file.write(str(py_list))


def to_json(py_list: list, path: str) -> None:
    new_path = path[:path.rfind('.') + 1] + 'json'
    with open(new_path, 'w', encoding='utf-8') as file:
        file.seek(0)
        file.write(json.dumps(py_list))


def to_csv(py_list: list, path: str) -> None:
    new_path = path[:path.rfind('.') + 1] + 'csv'
    with open(new_path, 'w', encoding='utf-8', newline='') as file:
        file.seek(0)
        writer = csv.writer(file)
        writer.writerow(py_list[0].keys())
        [writer.writerow(element.values()) for element in py_list]


def to_xlsx(py_list: list, path: str) -> None:
    new_path = path[:path.rfind('.') + 1] + 'xlsx'
    wb = openpyxl.Workbook(new_path)
    try:
        wb.save(new_path)

        wb = openpyxl.load_workbook(new_path)
        name_sheet = input('Type the name of the sheet to write the data in: ')
        sheet = wb.create_sheet(name_sheet)
        headers = list(py_list[0].keys())
        for row in sheet.iter_rows(min_row=1, max_col=len(headers), max_row=1):
            index = 0
            for cell in row:
                cell.value = headers[index]
                index += 1

        index_list = 0
        for row in sheet.iter_rows(min_row=2, max_col=len(headers), max_row=len(py_list) + 1):
            index_head = 0
            for cell in row:
                cell.value = py_list[index_list][headers[index_head]]
                index_head += 1
            index_list += 1
        wb.save(new_path)
    except PermissionError:
        print(f'Close file {new_path} to get permission to change it.')
        raise PermissionError


@decorator
def change_format() -> None:
    """
    Collects data from .txt, .json, .csv or .xlxs file,
    converts it into pythonic list with dicts,
    then creates new file of chosen type with this data.
    """
    source_path = input('Type the path to and the name of the source file: ')
    from_type = source_path[source_path.rfind('.') + 1:]

    print("Which new format do you choose? Type the number from 1 to 4:")
    to_type_num = input('1. txt\n'
                        '2. json\n'
                        '3. csv\n'
                        '4. xlsx: ')
    index_types = {'1': 'txt', '2': 'json', '3': 'csv', '4': 'xlsx'}

    if to_type_num not in ['1', '2', '3', '4']:
        print('There is no option for this number.')
        return

    elif from_type == index_types[to_type_num]:
        print('You have chosen the same format you already have.')
        return

    else:
        index_functions_to = {'1': to_txt, '2': to_json, '3': to_csv, '4': to_xlsx}
        functions_from = {'txt': from_txt, 'json': from_json, 'csv': from_csv, 'xlsx': from_xlsx}

        py_data = functions_from[from_type](source_path)  # in py_data should be data in form of a list with dicts

        index_functions_to[to_type_num](py_data, source_path)
