"""
    **This submodule  contain the class and functions used to read and write the LINDA excel documents.**
"""
import numpy as np
from openpyxl import load_workbook
from Excelacceslib import logger

logger = logger.getChild(__name__)


class Excel_Operate:
    """
        This clase is used to aperate with excel docs.
    """

    def __init__(self, filepath):
        """
        Initialize Excel_Operate object.
        :param filepath: (string) Absolute path to excel document.
        """
        try:
            self.path = filepath
            # data_only ignores loading formulas and instead loads only the resulting values.
            self.workbook = load_workbook(filename=filepath, data_only=True)
            self.sheetnames = self.workbook.sheetnames
            # self.shell.info(self.sheetnames)
        except FileNotFoundError:
            logger.error(f"Path error, filen not found: '{filepath}'")
            raise FileNotFoundError

    def save(self):
        """
        Initialize Excel_Operate object.
        :return: (int) If correct return 1
        """
        try:
            self.workbook.save(self.path)
            return None
        except:
            return 1

    def rangeval_to_matrix(self, _ranges, sheet_name):
        """
        Used for read chip register table.
        :param _ranges: (Array of String) Excel cells to get, Ex: ["B2:AE20", ]
        :param sheet_name: (String) Excel sheet name.
        :return: (Array) Chip register matrix.
        """
        sheet = self.workbook[sheet_name]
        # self.shell.info(f"Getting values of: {sheet.title}")

        tables = []
        for _range in _ranges:
            # Read table values
            values = sheet[_range]
            container = np.zeros((np.shape(values)))  # Create matrix with equal size as values readed
            row = 0
            for row_vals in values:
                column = 0
                for val in row_vals:
                    container[row][column] = val.value
                    column += 1
                row += 1
            tables.append(container)

        return np.array(tables, dtype=np.float64)

    def get_n_sheet_matrix(self, _ranges, sheet_pos):
        """
        Used for read pixel register table.
        :param _ranges: (Array of String) Excel cells to get, Ex: ["B2:AE20", "B20:AE20"]
        :param sheet_pos: (Matrix) The positions in sheet where the info is placed.
        :return: (Array) Pixel register matrix.
        """
        mx = []
        if sheet_pos == "all":
            i = 0
            for sheet_name in self.sheetnames:  # Getiing all sheets except the first one.
                if not i == 0:
                    mx.append(self.rangeval_to_matrix(_ranges, sheet_name))
                i += 1

        else:
            sheet_name = self.sheetnames[sheet_pos]
            mx.append(self.rangeval_to_matrix(_ranges, sheet_name))

        return np.array(mx)

    def matrix_to_rangeval(self, _ranges, sheet_name, matrix, _bool):
        """
        Used for write chip register and pixel regsiter table.
        :param _ranges: (Array of String) Excel cells to get, Ex: ["B2:AE20", ]
        :param sheet_name: (String) Excel sheet name.
        :param matrix: (Matrix) Chip Register matrix to write.
        :param _bool: (Bool) True if the data write is BOOLEAN data.
        """
        sheet = self.workbook[sheet_name]
        chip_iterate = 0
        for _range in _ranges:
            row_iterate = 0
            for row in sheet.iter_rows(min_row=_range[0], max_row=_range[1], min_col=_range[2], max_col=_range[3]):
                column_iterate = 0
                if len(matrix) == 1:  # The code is analyzing the chip register matrix
                    if row_iterate == 11 or row_iterate == 12 or row_iterate == 17 or row_iterate == 18:
                        _bool = True
                    else:
                        _bool = False

                for val in row:
                    if _bool:
                        val.value = bool(matrix[chip_iterate][row_iterate][column_iterate])
                    else:
                        val.value = matrix[chip_iterate][row_iterate][column_iterate]
                    column_iterate += 1
                row_iterate += 1
            chip_iterate += 1

    def send_pixelreg_matrix(self, _ranges, matrix):
        """
        Used for write pixel regsiter table.
        :param _ranges: (Array of String) Excel cells to get, Ex: ["B2:AE20", ]
        :param matrix: (Matrix) Chip Register matrix to write.
        """
        _bool = False
        i = 0
        for sheet_name in self.sheetnames[1:]:  # Geting all sheets except the first one.
            # Add all the sheet positions where dtype are boolean
            if i == 0 or i == 1 or i == 2 \
                    or i == 4 or i == 5 or i == 6 \
                    or i == 8 or i == 9 or i == 10 \
                    or i == 12 or i == 13 or i == 14 \
                    or i == 16 or i == 17 or i == 18 \
                    or i == 20 or i == 21 or i == 22 \
                    or i == 24 or i == 25 or i == 26 \
                    or i == 28 or i == 29 or i == 30 \
                    or i == 32 or i == 33 or i == 34 \
                    or i == 35 or i == 36 or i == 41 or i == 42 or i == 43:
                _bool = True
            else:
                _bool = False

            self.matrix_to_rangeval(_ranges, sheet_name, matrix[i], _bool)
            i += 1


def get_linda_matrix(path):
    """
    Get data from LINDA excel.
    :param filepath: (string) Absolute path to excel document.
    :return: (matrix) Retrun chip register matrix and pixel register matrix
    """
    # 0.0349 to execute
    workbook = Excel_Operate(path)

    # Getting chip register matrix
    chipe_reg_range = ["B2:AE20", ]
    sheet_name = "ChipReg"

    chip_reg = workbook.rangeval_to_matrix(chipe_reg_range, sheet_name)
    logger.info(np.shape(chip_reg))

    # Getting pixel register matrix
    pixel_reg_ranges = ["C6:V13", "Y6:AR13", "C16:V23", "Y16:AR23", "C26:V33", "Y26:AR33",
                        "C36:V43", "Y36:AR43", "C46:V53", "Y46:AR53", "C56:V63", "Y56:AR63",
                        "C66:V73", "Y66:AR73", "C76:V83", "Y76:AR83", "C86:V93", "Y86:AR93",
                        "C96:V103", "Y96:AR103", "C106:V113", "Y106:AR113", "C116:V123", "Y116:AR123",
                        "C126:V133", "Y126:AR133", "C136:V143", "Y136:AR143", "C146:V153", "Y146:AR153"]
    sheet_pos_pixel_reg = "all"

    pixel_reg = workbook.get_n_sheet_matrix(pixel_reg_ranges, sheet_pos_pixel_reg)
    logger.info(np.shape(pixel_reg))

    return chip_reg, pixel_reg


def write_linda_matrix(path, chip_reg, pixel_reg):
    """
    Write data to LINDA excel.
    :param filepath: (string) Absolute path to excel document.
    :param chip_reg: (matrix) Chip register data.
    :param pixel_reg: (matrix) Pixel register data.
    :return: (matrix) True if error.
    """
    workbook = Excel_Operate(path)

    # Writting chip register matrix
    chipe_reg_range = [[2, 20, 2, 31], ]
    sheet_name = "ChipReg"
    workbook.matrix_to_rangeval(chipe_reg_range, sheet_name, chip_reg, False)  # I will use for C

    # Writting pixel register matrix
    pixel_reg_ranges = [[6, 13, 3, 22], [6, 13, 25, 44], [16, 23, 3, 22], [16, 23, 25, 44], [26, 33, 3, 22],
                        [26, 33, 25, 44]
        , [36, 43, 3, 22], [36, 43, 25, 44], [46, 53, 3, 22], [46, 53, 25, 44], [56, 63, 3, 22], [56, 63, 25, 44]
        , [66, 73, 3, 22], [66, 73, 25, 44], [76, 83, 3, 22], [76, 83, 25, 44], [86, 93, 3, 22], [86, 93, 25, 44]
        , [96, 103, 3, 22], [96, 103, 25, 44], [106, 113, 3, 22], [106, 113, 25, 44]
        , [116, 123, 3, 22], [116, 123, 25, 44], [126, 133, 3, 22], [126, 133, 25, 44], [136, 143, 3, 22]
        , [136, 143, 25, 44], [146, 153, 3, 22], [146, 153, 25, 44]]

    workbook.send_pixelreg_matrix(pixel_reg_ranges, pixel_reg)

    error = workbook.save()
    if error:
        logger.error("Can't save file, check if xlsx is opened.")
        return True
    else:
        return False
